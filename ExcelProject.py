import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']


for row in range(2, sheet.max_row+1):
    cell = sheet.cell(row, 3)
    corrected_prices = (cell.value)*0.9
    sheet.cell(row, 4).value = corrected_prices


wb.save('transactions2.xlsx')
